#!/usr/bin/env python3
import json, os, requests
from datetime import datetime, timedelta

BOT_TOKEN = os.getenv("DORAYAID_BOT_TOKEN", "")
GROUP_ID = -5248748067
LOG_FILE = "/root/dorayaid_transfers.json"

def today_bkk():
    return (datetime.utcnow() + timedelta(hours=7)).strftime("%Y-%m-%d")

def yesterday_bkk():
    return (datetime.utcnow() + timedelta(hours=7) - timedelta(days=1)).strftime("%Y-%m-%d")

def send(text):
    requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
        json={"chat_id": GROUP_ID, "text": text, "parse_mode": "Markdown"}, timeout=10)

def save_monthly_excel():
    """บันทึก Excel รายเดือน"""
    import openpyxl
    if not os.path.exists(LOG_FILE): return
    db = json.load(open(LOG_FILE))
    yesterday = yesterday_bkk()
    month = yesterday[:7]  # YYYY-MM
    excel_dir = "/root/dorayaid_excel"
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = f"{excel_dir}/slip_{month}.xlsx"
    
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = month
        ws.append(["#", "วันที่", "เวลา", "ผู้ส่ง", "ยอด (บาท)", "ผู้โอน", "ผู้รับ"])
    
    ws = wb.active
    trans = [t for t in db.get("transfers", []) if t["date"].startswith(yesterday)]
    row_num = ws.max_row
    for i, t in enumerate(trans, 1):
        ws.append([row_num + i - 1, t.get("date",""), t.get("slip_time",""), t.get("sender",""), t.get("amount",0), t.get("bank",""), t.get("receiver","")])
    
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)
    
    wb.save(excel_path)
    
    # ส่ง Excel เข้ากลุ่ม
    with open(excel_path, "rb") as f:
        requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
            data={"chat_id": GROUP_ID, "caption": f"📁 Excel ประจำเดือน {month}"},
            files={"document": (f"slip_{month}.xlsx", f)}, timeout=15)

if os.path.exists(LOG_FILE):
    db = json.load(open(LOG_FILE))
    yesterday = yesterday_bkk()
    trans = [t for t in db.get("transfers", []) if t["date"].startswith(yesterday)]
    if trans:
        total = sum(t["amount"] for t in trans)
        txt = f"📊 *สรุปยอดโอนวันที่ {yesterday}*\n\n"
        for i, t in enumerate(trans, 1):
            txt += f"{i}. {t['sender']} — *{t['amount']:,.2f}* บาท"
            if t.get("receiver"): txt += f" → {t['receiver']}"
            if t.get("slip_time"): txt += f" 🕐 {t['slip_time']}"
            txt += "\n"
        txt += f"\n💰 *รวมทั้งหมด: {total:,.2f} บาท* ({len(trans)} รายการ)"
        send(txt)
    else:
        send(f"📊 สรุปวันที่ {yesterday}\nไม่มีรายการโอนค่ะ")

save_monthly_excel()
